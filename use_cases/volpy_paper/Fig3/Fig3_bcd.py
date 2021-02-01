#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Fri Sep 20 10:33:04 2019
@author: Changjia
This file produces Fig3_bcd, Table 1,2,3 and Fig S4 for volpy paper.
"""
#%%
import caiman as cm
import os
from caiman.base.rois import nf_read_roi_zip
import numpy as np
from caiman.source_extraction.cnmf.estimates import Estimates, compare_components
from caiman.base.rois import nf_match_neurons_in_binary_masks
import skimage
import matplotlib
matplotlib.rcParams['pdf.fonttype'] = 42
matplotlib.rcParams['ps.fonttype'] = 42
import matplotlib.pyplot as plt
try:
    if __IPYTHON__:
        # this is used for debugging purposes only. allows to reload classes
        # when changed
        get_ipython().magic('load_ext autoreload')
        get_ipython().magic('autoreload 2')
except NameError:
    pass

#%% fig3_b F1 scores for all datasets
folder = '/home/nel/Code/NEL_LAB/Mask_RCNN/result_f1/VolPy/cross'
files = os.listdir(folder)
val = {} 
for file in files:
    if 'val' in file:
        fnames = folder + '/' + file
        data = np.load(fnames, allow_pickle=True).item()
        for keys, values in data.items():
            val[keys] = values['f1_score']

plt.figure(figsize=(15,15))
keys = ['403106_3min', 'FOV1', 'FOV1_35um', 'FOV2', 'FOV2_80um', 'FOV3',
 'FOV3_35um', 'FOV4', 'FOV4_50um', '06152017Fish1-2', '10192017Fish2-1',
 '10192017Fish3-1', 'IVQ29_S5_FOV4', 'IVQ29_S5_FOV6', 'IVQ32_S2_FOV1',
 'IVQ38_S1_FOV5', 'IVQ38_S2_FOV3', 'IVQ39_S1_FOV7', 'IVQ39_S2_FOV3',
 'IVQ39_S2_FOV4', 'IVQ48_S7_FOV1', 'IVQ48_S7_FOV5', 'IVQ48_S7_FOV7', 'IVQ48_S7_FOV8']
              
values = [val[key] for key in keys]
i = 0
a = np.arange(0,9)
plt.bar(a,values[a[0]:a[-1]+1], width=0.5, label='L1', color='red')
b= np.arange(9,12)
plt.bar(b,values[b[0]:b[-1]+1], width=0.5, label='TEG', color='blue')
c= np.arange(12,24)
plt.bar(c,values[c[0]:], width=0.5, label='HPC', color='gray')
plt.legend()
plt.ylabel('F1 scores')
plt.xticks(np.arange(len(keys)), keys, rotation='vertical',fontsize=8)
plt.legend()
ax = plt.gca()
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)
plt.tight_layout()
#plt.savefig('/home/nel/NEL-LAB Dropbox/NEL/Papers/VolPy/Figures/plos/original_files/figure3/F1_all.pdf')

file_names = os.listdir('/home/nel/NEL-LAB Dropbox/NEL/Papers/VolPy/zenodo')
file_names = sorted([file[:-4] for file in file_names if '.txt' not in file])
file_names = ['L1.00.00', 'L1.01.00', 'L1.01.35', 'L1.02.00', 'L1.02.80', 'L1.03.00',
              'L1.03.35', 'L1.04.00', 'L1.04.50', 'TEG.01.02', 'TEG.02.01','TEG.03.01',
              'HPC.29.04', 'HPC.29.06', 'HPC.32.01', 'HPC.38.05', 'HPC.38.03', 'HPC.39.07',
              'HPC.39.03', 'HPC.39.04', 'HPC.48.01', 'HPC.48.05', 'HPC.48.07', 'HPC.48.08']

#%%
excel_folder = '/home/nel/NEL-LAB Dropbox/NEL/Papers/VolPy/Figures/plos/excel_data'
df = pd.DataFrame({'file':file_names,'f1 score':values})
dfs = [df]
text = 'F1 score of VolPy for all evaluated datasets in validation.'
fig_name = 'Fig 3b'
excel_name = os.path.join(excel_folder, 'volpy_data.xlsx')
# run function
multiple_dfs(dfs, fig_name, excel_name, 2, text)

#%% fig3_c and table 3. F1 score for cross validation results
folder = '/home/nel/Code/NEL_LAB/Mask_RCNN/result_f1/VolPy/cross'
files = os.listdir(folder)
results = {}    
metric = ['f1_score', 'precision', 'recall'][0]
summary = {}
for mode in ['train','val']:
    result = {'L1':[], 'Fish':[],'IVQ':[]}
    for file in files:
        if mode in file:
            fnames = folder + '/' + file
            data = np.load(fnames, allow_pickle=True).item()
            fish = []; L1 = []; IVQ = [];
            for keys, values in data.items():
                if 'Fish' in keys:
                    fish.append(values[metric])
                elif 'IVQ' in keys:
                    IVQ.append(values[metric])
                else:
                    L1.append(values[metric])
            result['Fish'].append(np.array(fish).mean())
            result['IVQ'].append(np.array(IVQ).mean())
            result['L1'].append(np.array(L1).mean())

    results[mode] = result       
    summary[mode] = {'L1':[np.array(result['L1']).mean(), np.std(np.array(result['L1']))], 
           'Fish':[np.array(result['Fish']).mean(), np.std(np.array(result['Fish']))], 
           'IVQ':[np.array(result['IVQ']).mean(), np.std(np.array(result['IVQ']))]}
print(summary)
# Manual annotations results
manual = np.load('/home/nel/Code/NEL_LAB/Mask_RCNN/result_f1/VolPy/manual/voltage_v1.2_manual.npy', allow_pickle=True).item() 

result = {'L1':[], 'Fish':[],'IVQ':[]}
for keys, values in manual.items():
    result['L1'].append(values['f1_score']['L1'])
    result['Fish'].append(values['f1_score']['TEG'])
    result['IVQ'].append(values['f1_score']['HPC'])
results['manual'] = result
    
print(np.array(result['L1']).mean())
print(np.array(result['Fish']).mean())
print(np.array(result['IVQ']).mean())
print(np.array(result['L1']).std())
print(np.array(result['Fish']).std())
print(np.array(result['IVQ']).std())

# Bar plot
t_mean = [np.array(i).mean() for i in results['train'].values()]
v_mean = [np.array(i).mean() for i in results['val'].values()]
m_mean = [np.array(i).mean() for i in results['manual'].values()]
t_std = [np.std(np.array(i)) for i in results['train'].values()]
v_std = [np.std(np.array(i)) for i in results['val'].values()]
m_std = [np.std(np.array(i)) for i in results['manual'].values()]

labels=['L1','TEG','HPC']
F1_mean = np.stack([t_mean,v_mean, m_mean], axis=1)
F1_std = np.stack([t_std, v_std, m_std], axis=1)
x = np.arange(len(labels))
width = 0.2       # the width of the bars: can also be len(x) sequence

plt.figure()
plt.title('F1 score')
ax = plt.gca()
rects1 = ax.bar(x - width, F1_mean[:,0], width, yerr=F1_std[:,0], label='train', error_kw=dict(ecolor='gray', lw=1.5, capsize=2, capthick=1))
rects2 = ax.bar(x, F1_mean[:,1], width, yerr=F1_std[:,1], label='val', error_kw=dict(ecolor='gray', lw=1.5, capsize=2, capthick=1))
rects2 = ax.bar(x + width, F1_mean[:,2], width, yerr=F1_std[:,2], label='manual', error_kw=dict(ecolor='gray', lw=1.5, capsize=2, capthick=1))


# Add some text for labels, title and custom x-axis tick labels, etc.
ax.set_ylabel('F1 Score')
ax.set_xticks(x)
ax.set_xticklabels(labels)
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)
ax.legend()
#plt.savefig('/home/nel/NEL-LAB Dropbox/NEL/Papers/VolPy/Figures/plos/original_files/figure3/F1_average.pdf')

#%%
df1 = pd.DataFrame({})
for s1 in results.keys():
    for idx, s2 in enumerate(labels):
        df1[s1 + '_' + s2] = results[s1][list(results['train'].keys())[idx]]

#%%
df2 = pd.DataFrame({'file':['L1', 'TEG', 'HPC'],'train_mean':t_mean, 'val_mean':v_mean, 'human_mean':m_mean, 
                   'train_std':t_std, 'val_std':v_std, 'human_std':m_std})

        
#%%
import openpyxl
def multiple_dfs(df_list, sheets, file_name, spaces, text):
    try:
        book = load_workbook(file_name)
        print('existing workbook')
    except:
        book = openpyxl.Workbook()
        print('new workbook')

    writer = pd.ExcelWriter(file_name,engine='openpyxl') 
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    row = 2
    for dataframe in df_list:
        dataframe.to_excel(writer,sheet_name=sheets,startrow=row, startcol=0, index=False, na_rep='NA')   
        row = row + len(dataframe.index) + spaces + 1
    #writer.sheets[sheets].cell(1,1).style.alignment.wrap_text = True
    writer.sheets[sheets].cell(1,1).value = text
    writer.save()


# list of dataframes
dfs = [df1,df2]
text = 'Average F1 score on training and validation sets grouped by dataset type. Results were provided for training, validation and human annotators (against consensus ground truth)'
fig_name = 'Fig 3c'
excel_name = os.path.join(excel_folder, 'volpy_data.xlsx')
# run function
multiple_dfs(df_list=dfs, sheets=fig_name, file_name=excel_name, spaces=2, text=text)

#%% fig 3_d F1 score with different number of neurons
folder = '/home/nel/Code/NEL_LAB/Mask_RCNN/result_f1/VolPy/40_epochs'
files = sorted(os.listdir(folder))
F1_all = {'train':{}, 'val':{}}
modes = ['train', 'val']
for file in files:
    F1 = []
    for mode in modes:
        if mode in file:
            #F1[mode] = []
            path = os.path.join(folder, file) 
            data = np.load(path, allow_pickle=True).item()
            li = []
            for keys, values in data.items():
                li.append(values['f1_score'])
            F1.append(np.array(li).mean())
            F1_all[mode][file] = F1
labels = {'L1':['29', '65', '115', '217', '333'], 'TEG':['28', '69'], 'HPC':['21', '41']}
results = {'train':[], 'val':[]}
for name in ['L1', 'TEG', 'HPC']:
    result = {'train':[], 'val':[]}
    plt.figure()
    ax = plt.gca()
    width = 0.35
    x = np.arange(len(labels[name]))
    for keys, values in F1_all['train'].items():
        if name in keys:
            result['train'].append(F1_all['train'][keys][0])
    for keys, values in F1_all['val'].items():
        if name in keys:
            result['val'].append(F1_all['val'][keys][0])        
    results['train'] = results['train'] + result['train']
    results['val'] = results['val'] + result['val']
    rects1 = ax.bar(x - width/2, result['train'], width, label='train')
    rects2 = ax.bar(x + width/2, result['val'], width, label='val')
    
    # Add some text for labels, title and custom x-axis tick labels, etc.
    ax.set_ylabel('F1 Score')
    ax.set_xticks(x)
    ax.set_xticklabels(labels[name])
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    ax.legend()
    #plt.savefig(f'/home/nel/NEL-LAB Dropbox/NEL/Papers/VolPy/Figures/plos/original_files/figure3/F1_{name}.pdf')    
    #plt.close()

# number of neurons in each case
"""
folders = sorted(['/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_L1_0.5', '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_L1_1', '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_L1_2', '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_L1_4', '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_L1_6'])
folders = sorted(['/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_TEG_1', '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_TEG_2'])
folders = sorted(['/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_HPC_4_2', '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_HPC_8'])

for folder in folders:
    folder = os.path.join(folder, 'train')
    files = os.listdir(folder)
    files = [file for file in files if 'mask' in file]    
    n = 0
    for file in files:
        path = os.path.join(folder, file)
    
        m = np.load(path, allow_pickle=True)['mask']
        n = n + m.shape[0]
    print(n)
"""

#%%
num = ['29', '65', '115', '217', '333', '28', '69', '21', '41']
df = pd.DataFrame({'num of neurons for training': num,'train':results['train'], 'val':results['val']})
excel_folder = '/home/nel/NEL-LAB Dropbox/NEL/Papers/VolPy/Figures/plos/excel_data'
dfs = [df]
text = 'Performance of the network in function of training set size for each dataset type'
fig_name = 'Fig 3d'
excel_name = os.path.join(excel_folder, 'volpy_data.xlsx')
# run function
multiple_dfs(dfs, fig_name, excel_name, 2, text)

#%% Table 1 number of neurons for each type of datasets
folder = '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_all'
files = os.listdir(folder)
files = sorted([file for file in files if 'mask' in file])
L1 = 0
TEG = 0
HPC = 0
for file in files:
    path = os.path.join(folder, file)
    mm = np.load(path, allow_pickle=True)['mask']
    if 'Fish' in file:
        TEG = TEG + len(mm)
    elif 'IVQ' in file:
        HPC = HPC + len(mm)
    else:
        L1 = L1 + len(mm)
        
#%% Table 2 number of neurons for each dataset
folder = '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_all'
files = os.listdir(folder)
files = sorted([file for file in files if 'mask' in file])

for file in files:
    path = os.path.join(folder, file)
    mm = np.load(path, allow_pickle=True)['mask']
    print(f'{file}: {len(mm)}')
    
#%% Table 3 number of neurons
summary = {}
modes = ['train', 'val']
folders = ['/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2', 
          '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_cross2', 
          '/home/nel/Code/NEL_LAB/Mask_RCNN/datasets/voltage_v1.2_cross3']

for mode in modes:
    L1_all = []
    TEG_all = []
    HPC_all = []
    for folder in folders:
        files = os.listdir(os.path.join(folder, mode))
        files = sorted([file for file in files if 'mask' in file])
        L1 = 0
        TEG = 0
        HPC = 0
        for file in files:
            path = os.path.join(folder,mode, file)
            mm = np.load(path, allow_pickle=True)['mask']
            if 'Fish' in file:
                TEG = TEG + len(mm)
            elif 'IVQ' in file:
                HPC = HPC + len(mm)
            else:
                L1 = L1 + len(mm)
        L1_all.append(L1)
        TEG_all.append(TEG)
        HPC_all.append(HPC)
    summary[mode] = {'L1':[np.array(L1_all).mean(), np.std(np.array(L1_all))], 
           'Fish':[np.array(TEG_all).mean(), np.std(np.array(TEG_all))], 
           'IVQ':[np.array(HPC_all).mean(), np.std(np.array(HPC_all))]}


    
#%% This produces Fig S4
import tensorflow as tf
dataset_name = ["voltage_v1.2", "voltage_v1.2_cross2", "voltage_v1.2_cross3",
                "voltage_v1.2_L1_6", "voltage_v1.2_L1_4", "voltage_v1.2_L1_2", "voltage_v1.2_L1_1", 
                "voltage_v1.2_TEG_2", "voltage_v1.2_TEG_1", "voltage_v1.2_HPC_8", "voltage_v1.2_HPC_4",
                "voltage_v1.2_HPC_2", "voltage_v1.2_HPC_1", "voltage_v1.2_rerun", "voltage_v1.2_HPC_4_2", 
                'voltage_v1.2_L1_0.5']
weights = ["/neurons20200824T1032/mask_rcnn_neurons_0040.h5",
           "/neurons20200825T0951/mask_rcnn_neurons_0040.h5", 
           "/neurons20200825T1039/mask_rcnn_neurons_0040.h5",
           '/neurons20200901T0906/mask_rcnn_neurons_0030.h5',
           '/neurons20200901T1008/mask_rcnn_neurons_0030.h5',
           "/neurons20201116T1141/mask_rcnn_neurons_0030.h5", #'/neurons20200901T1058/mask_rcnn_neurons_0030.h5',
           '/neurons20200902T1530/mask_rcnn_neurons_0030.h5',
           "/neurons20200903T1124/mask_rcnn_neurons_0030.h5",
           "/neurons20200903T1215/mask_rcnn_neurons_0030.h5",
           '/neurons20200926T0919/mask_rcnn_neurons_0030.h5', 
           "/neurons20200926T1036/mask_rcnn_neurons_0030.h5",
           "/neurons20200926T1124/mask_rcnn_neurons_0030.h5", 
           "/neurons20200926T1213/mask_rcnn_neurons_0030.h5",
           "/neurons20201010T1758/mask_rcnn_neurons_0040.h5",
           "/neurons20201015T1403/mask_rcnn_neurons_0030.h5", 
           "/neurons20201019T1034/mask_rcnn_neurons_0030.h5"]

folders = [os.path.split(name)[0][1:] for name in weights]

loss = {}
logs_dir = '/home/nel/Code/NEL_LAB/Mask_RCNN/logs/'
#folder = 'neurons20201015T1403'
for idx, folder in enumerate(folders):
    train_loss = []
    val_loss = []
    loss[dataset_name[idx]] = {}
    files = os.listdir(os.path.join(logs_dir, folder))
    files = sorted([file for file in files if 'events' in file])
    for file in files:
        path = os.path.join(logs_dir, folder, file)
        for event in tf.train.summary_iterator(path):
            for value in event.summary.value:
                #print(value.tag)
                if value.tag == 'loss':
                    train_loss.append(value.simple_value)
                elif value.tag == 'val_loss':
                    val_loss.append(value.simple_value)
    
    loss[dataset_name[idx]]['loss'] = train_loss
    loss[dataset_name[idx]]['val_loss'] = val_loss



keys = ["voltage_v1.2", "voltage_v1.2_L1_6", "voltage_v1.2_L1_4", "voltage_v1.2_L1_2", 
                "voltage_v1.2_L1_1", "voltage_v1.2_L1_0.5", "voltage_v1.2_TEG_2", "voltage_v1.2_TEG_1", 
                "voltage_v1.2_HPC_8", "voltage_v1.2_HPC_4_2"]
loss_new = {}
for key in keys:
    loss_new[key] = loss[key]

def smooth(scalars, weight):  # Weight between 0 and 1
    last = scalars[0]  # First value in the plot (first timestep)
    smoothed = list()
    for point in scalars:
        smoothed_val = last * weight + (1 - weight) * point  # Calculate smoothed value
        smoothed.append(smoothed_val)                        # Save it
        last = smoothed_val                                  # Anchor the last smoothed value

    return smoothed

fig, axs = plt.subplots(4, 3)
i = 0
smooth_weight = 0.7
for keys, values in loss_new.items():
    h = int(np.floor(i/3))
    w = int(i - h * 3)
    train_loss = smooth(loss[keys]['loss'], smooth_weight)
    val_loss = smooth(loss[keys]['val_loss'], smooth_weight)
    x = list(range(1, len(train_loss)+1))
    axs[h, w].plot(x, train_loss)
    axs[h, w].plot(x, val_loss)
    keys = 'voltage_' + keys[13:]
    axs[h, w].title.set_text(keys)
    axs[h, w].spines['right'].set_visible(False)
    axs[h, w].spines['top'].set_visible(False)  
    #axs[h, w].spines['bottom'].set_visible(False)  
    axs[h, w].spines['left'].set_visible(True) 
    if h + w == 0:
        axs[h, w].legend(['train_loss', 'val_loss'])
        axs[h, w].spines['bottom'].set_visible(True)  
    i = i + 1
    plt.tight_layout()
    
#plt.savefig('/home/nel/NEL-LAB Dropbox/NEL/Papers/VolPy/Figures/plos/original_files/supp/loss_v2.0.pdf')
       